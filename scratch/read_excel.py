import openpyxl
from pathlib import Path

file_path = Path("old-backups/Motor List.xlsx")
if not file_path.exists():
    print("File not found:", file_path)
    exit(1)

wb = openpyxl.load_workbook(file_path, read_only=True)
print("Sheet names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\nSheet: {sheetname}")
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("  Empty sheet")
        continue
    print(f"  Total rows: {len(rows)}")
    print(f"  Headers: {rows[0]}")
    if len(rows) > 1:
        print(f"  Sample row 1: {rows[1]}")
    if len(rows) > 2:
        print(f"  Sample row 2: {rows[2]}")
