import openpyxl

wb = openpyxl.load_workbook(r"d:\motor data\MN3508 - Dataset Template.xlsx")
print("Sheet names:", wb.sheetnames)
sheet = wb.active
print("Active sheet title:", sheet.title)

for r in range(1, 20):
    row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
    if any(v is not None for v in row_vals):
        print(f"Row {r}: {row_vals}")
