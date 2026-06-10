import urllib.request
import json
import os
import openpyxl
from datetime import datetime

def load_env(env_path=".env"):
    env_data = {}
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    if '=' in line:
                        key, val = line.split('=', 1)
                        env_data[key.strip()] = val.strip()
    return env_data

env = load_env()
SUPABASE_URL = env.get("SUPABASE_URL")
SERVICE_ROLE_KEY = env.get("service_role")

def supabase_request(table, query_params=""):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if query_params:
        url += f"?{query_params}"
    
    headers = {
        "apikey": SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
        "Content-Type": "application/json"
    }
    
    req = urllib.request.Request(url, headers=headers, method="GET")
    try:
        with urllib.request.urlopen(req) as res:
            return json.loads(res.read().decode('utf-8'))
    except Exception as e:
        print(f"Error fetching {table}: {e}")
        raise e

def formatDate(date_str):
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        return dt.strftime("%d-%m-%Y")
    except Exception:
        return datetime.now().strftime("%d-%m-%Y")

def getExcelSheetName(motor_name):
    if not motor_name or motor_name == 'Draft Runs':
        return 'Draft Runs'
    import re
    name = re.sub(r'KV', '', motor_name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', ' ', name).strip()
    name = re.sub(r'[\\/\?\*\[\]]', '', name)
    return name[:31] or 'Sheet1'

def run_simulation_and_verification():
    motors = supabase_request("motors")
    motor_map = {m["id"]: m for m in motors}
    runs = supabase_request("motor_test_runs")
    pts = supabase_request("motor_test_data_points")

    print(f"Loaded {len(motors)} motors, {len(runs)} runs, {len(pts)} data points.")

    runs_data = {}
    for p in pts:
        run_id = p["test_run_id"]
        if run_id not in runs_data:
            run_meta = next((r for r in runs if r["id"] == run_id), None)
            if run_meta:
                runs_data[run_id] = {
                    "metadata": run_meta,
                    "dataPoints": []
                }
        if run_id in runs_data:
            runs_data[run_id]["dataPoints"].append(p)

    for r_id in runs_data:
        runs_data[r_id]["dataPoints"].sort(key=lambda x: float(x["throttle"]) if x["throttle"] is not None else 0.0)

    motors_runs = {}
    for r_id, run_obj in runs_data.items():
        m_id = run_obj["metadata"]["motor_id"]
        motor = motor_map.get(m_id)
        motor_name = motor["motor_name"] if motor else "Draft Runs"
        if motor_name not in motors_runs:
            motors_runs[motor_name] = []
        motors_runs[motor_name].append(run_obj)

    for m_name in motors_runs:
        motors_runs[m_name].sort(key=lambda x: x["metadata"]["tested_at"])

    sim_wb = openpyxl.Workbook()
    default_sheet = sim_wb.active
    sim_wb.remove(default_sheet)

    for m_name, runs_list in motors_runs.items():
        sheet_name = getExcelSheetName(m_name)
        sheet = sim_wb.create_sheet(title=sheet_name)
        
        row_idx = 1
        for run_idx, run_obj in enumerate(runs_list):
            if run_idx > 0:
                row_idx += 2

            meta = run_obj["metadata"]
            pts_list = run_obj["dataPoints"]

            sheet.cell(row=row_idx, column=1, value='# Software name: created by ROTRIX')
            sheet.cell(row=row_idx, column=2, value=f"Device Name: {meta.get('device_name') or 'test device'}")
            sheet.cell(row=row_idx, column=3, value=f"Motor Model: {m_name}")
            sheet.cell(row=row_idx, column=4, value=f"Propeller Model:{meta.get('propeller_model') or ''}")
            row_idx += 1

            sheet.cell(row=row_idx, column=1, value=f"# Test conducted by: {meta.get('test_conducted_by') or 'Unknown'}")
            sheet.cell(row=row_idx, column=2, value=f"Powertrain Name: {meta.get('powertrain_name') or 'test prowertrain'}")
            sheet.cell(row=row_idx, column=3, value=f"ESC Model: {meta.get('esc_model') or ''}")
            sheet.cell(row=row_idx, column=4, value=f"Battery Voltage and Capacity: {meta.get('battery_voltage_capacity') or 'None'}")
            sheet.cell(row=row_idx, column=5, value=f"Battery : {meta.get('battery_info') or ''}")
            row_idx += 1

            sheet.cell(row=row_idx, column=1, value=f"# Generated On: {formatDate(meta.get('tested_at'))}")
            row_idx += 1
            row_idx += 1

            headers = ['Throttle', 'Voltage\n(V)', 'Current\n(A)', 'Power\n(W)', 'Thrust\n(G)', 'RPM', 'Efficiency\n(G/W)', 'Temperature\n(℃)']
            for c_idx, h in enumerate(headers, 1):
                sheet.cell(row=row_idx, column=c_idx, value=h)
            row_idx += 1

            for dp in pts_list:
                v = float(dp["voltage"]) if dp["voltage"] is not None else 0.0
                i = float(dp["current"]) if dp["current"] is not None else 0.0
                p = round(v * i, 2)
                t = float(dp["thrust_g"]) if dp["thrust_g"] is not None else 0.0
                eff = round(t / p, 2) if p > 0 else 0.0
                
                raw_throttle = float(dp["throttle"]) if dp["throttle"] is not None else 0.0
                throttle = raw_throttle / 100.0 if raw_throttle > 1.0 else raw_throttle

                row_vals = [
                    throttle,
                    dp["voltage"],
                    dp["current"],
                    p,
                    dp["thrust_g"],
                    dp["rpm"],
                    eff,
                    dp["temperature"]
                ]
                for c_idx, val in enumerate(row_vals, 1):
                    sheet.cell(row=row_idx, column=c_idx, value=val)
                row_idx += 1

    sim_path = r"d:\motor data\scratch\simulated_export.xlsx"
    sim_wb.save(sim_path)
    print(f"Simulated export saved to {sim_path}")

    tmpl_path = r"d:\motor data\template datasets\MN3508 - Dataset Template.xlsx"
    tmpl_wb = openpyxl.load_workbook(tmpl_path, data_only=True)

    print("\n--- Comparing simulated export vs template ---")
    print(f"Template sheets: {tmpl_wb.sheetnames}")
    print(f"Simulated sheets: {sim_wb.sheetnames}")

    for s_name in sim_wb.sheetnames:
        if s_name not in tmpl_wb.sheetnames:
            print(f"Sheet '{s_name}' is in simulated export but not in template.")
            continue

        sim_sheet = sim_wb[s_name]
        tmpl_sheet = tmpl_wb[s_name]

        max_rows = max(sim_sheet.max_row, tmpl_sheet.max_row)
        mismatches = 0

        for r in range(1, max_rows + 1):
            sim_row = [sim_sheet.cell(row=r, column=c).value for c in range(1, 9)]
            tmpl_row = [tmpl_sheet.cell(row=r, column=c).value for c in range(1, 9)]

            if sim_row != tmpl_row:
                is_mismatch = False
                for c_idx in range(8):
                    v_sim = sim_row[c_idx]
                    v_tmpl = tmpl_row[c_idx]
                    if v_sim == v_tmpl:
                        continue
                    
                    try:
                        if abs(float(v_sim) - float(v_tmpl)) < 0.02:
                            continue
                    except (ValueError, TypeError):
                        pass

                    if isinstance(v_sim, str) and isinstance(v_tmpl, str):
                        if v_sim.strip().lower() == v_tmpl.strip().lower():
                            continue
                        import re
                        s_norm = re.sub(r'\s+', ' ', v_sim.strip().lower())
                        t_norm = re.sub(r'\s+', ' ', v_tmpl.strip().lower())
                        if s_norm == t_norm:
                            continue

                    is_mismatch = True
                    break

                if is_mismatch:
                    mismatches += 1
                    # Only print first 5 mismatches per sheet to prevent log bloat
                    if mismatches <= 5:
                        print(f"[{s_name}] Mismatch Row {r:02d}:")
                        print(f"  Sim: {sim_row}")
                        print(f"  Tmp: {tmpl_row}")

        if mismatches == 0:
            print(f"✓ Sheet '{s_name}' matches template layout and values perfectly!")
        else:
            print(f"✗ Sheet '{s_name}' has {mismatches} mismatches total.")

if __name__ == "__main__":
    run_simulation_and_verification()
