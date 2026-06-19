# admin_console.py
import os
import ssl
import json
import uuid
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pg8000.dbapi
import boto3
import openpyxl
from datetime import datetime

# ── Load .env file ──────────────────────────────────────────────────────────
def load_env(path='.env'):
    env = {}
    if os.path.exists(path):
        for line in open(path, encoding='utf-8'):
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                val = v.strip()
                if (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
                    val = val[1:-1]
                env[k.strip()] = val
    return env

ENV = load_env()

# Database Configs
DB_HOST = ENV.get("DB_HOST", "database-2-instance-1.c768qwys6eaf.eu-north-1.rds.amazonaws.com")
DB_PORT = int(ENV.get("DB_PORT", "5432"))
DB_NAME = ENV.get("DB_NAME", "postgres")
DB_USER = ENV.get("DB_USER", "postgres")
DB_PASSWORD = ENV.get("DB_PASSWORD", "ThrustVault321")
USE_AWS_IAM_AUTH = ENV.get("USE_AWS_IAM_AUTH", "false").lower() in ['true', '1', 'yes']
AWS_REGION = ENV.get("AWS_REGION", "eu-north-1")

# Cognito Configs
COGNITO_USER_POOL_ID = ENV.get("COGNITO_USER_POOL_ID", "eu-north-1_8bwG74uX8")
COGNITO_CLIENT_ID = ENV.get("COGNITO_CLIENT_ID", "3e2o8v80oqjn18d3frdsvobrfb")
COGNITO_REGION = ENV.get("COGNITO_REGION", "eu-north-1")

# ── AWS Connection Helpers ───────────────────────────────────────────────────
def get_cognito_client():
    # Attempt to load custom credential profiles
    if 'AWS_PROFILE' not in os.environ and 'AWS_ACCESS_KEY_ID' not in os.environ:
        try:
            profiles = boto3.Session().available_profiles
            for p in ['ThrustVault', 'Bharani-Claude-api']:
                if p in profiles:
                    return boto3.Session(profile_name=p).client('cognito-idp', region_name=COGNITO_REGION)
        except Exception:
            pass
    return boto3.client('cognito-idp', region_name=COGNITO_REGION)

def get_db_pass():
    if USE_AWS_IAM_AUTH:
        try:
            rds_client = boto3.client('rds', region_name=AWS_REGION)
            return rds_client.generate_db_auth_token(
                DBHostname=DB_HOST,
                Port=DB_PORT,
                DBUsername=DB_USER,
                Region=AWS_REGION
            )
        except Exception as e:
            print(f"Warning: Failed to generate IAM DB Token: {e}. Using password.")
    return DB_PASSWORD

def get_db_connection():
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE
    return pg8000.dbapi.connect(
        host=DB_HOST,
        port=DB_PORT,
        database=DB_NAME,
        user=DB_USER,
        password=get_db_pass(),
        ssl_context=ssl_context,
        timeout=10.0
    )

# ── Tkinter Application ──────────────────────────────────────────────────────
class ThrustVaultAdminApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ThrustVault UAV Motor Admin Console")
        self.root.geometry("1100x700")
        
        # Configure overall layout grid
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        # Style System
        self.style = ttk.Style()
        self.style.theme_use("clam")
        
        # Color Palette Definitions
        self.BG_DARK = "#0f172a"
        self.BG_CARD = "#1e293b"
        self.BORDER_COLOR = "#334155"
        self.ACCENT_COLOR = "#3b82f6"
        self.TEXT_COLOR = "#f8fafc"
        self.TEXT_MUTED = "#94a3b8"

        self.root.configure(bg=self.BG_DARK)

        # Apply Custom Styles
        self.style.configure(".", background=self.BG_DARK, foreground=self.TEXT_COLOR, font=("Segoe UI", 10))
        self.style.configure("TLabel", background=self.BG_DARK, foreground=self.TEXT_COLOR)
        self.style.configure("TNotebook", background=self.BG_DARK, borderwidth=0)
        self.style.configure("TNotebook.Tab", background=self.BG_CARD, foreground=self.TEXT_MUTED, font=("Segoe UI", 10, "bold"), padding=(15, 6))
        self.style.map("TNotebook.Tab", background=[("selected", self.ACCENT_COLOR)], foreground=[("selected", "#ffffff")])
        
        self.style.configure("TFrame", background=self.BG_DARK)
        self.style.configure("Card.TFrame", background=self.BG_CARD, borderwidth=1, relief="solid")
        
        self.style.configure("TButton", background=self.ACCENT_COLOR, foreground="#ffffff", borderwidth=0, font=("Segoe UI", 9, "bold"), padding=6)
        self.style.map("TButton", background=[("active", "#2563eb")])
        self.style.configure("Danger.TButton", background="#ef4444")
        self.style.map("Danger.TButton", background=[("active", "#dc2626")])
        
        self.style.configure("Treeview", background=self.BG_CARD, foreground=self.TEXT_COLOR, fieldbackground=self.BG_CARD, rowheight=26, borderwidth=0)
        self.style.configure("Treeview.Heading", background=self.BORDER_COLOR, foreground=self.TEXT_COLOR, borderwidth=0, font=("Segoe UI", 10, "bold"))
        self.style.map("Treeview", background=[("selected", "#2563eb")], foreground=[("selected", "#ffffff")])

        # Main Layout
        self.notebook = ttk.Notebook(self.root)
        self.notebook.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        # Create Tabs
        self.create_motors_tab()
        self.create_categories_tab()
        self.create_users_tab()
        self.create_schema_tab()
        self.create_logs_tab()
        self.create_data_tab()

        # Connect & Load Initially
        self.log_event("SYSTEM", "Console started successfully.")

    def log_event(self, action, details):
        """Helper to post audit logs of operations performed in Tkinter admin console."""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT INTO public.audit_logs 
                (email, role, route, method, status, ip_address, user_agent, risk_level, details) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                ["DesktopConsoleAdmin", "admin", f"DESKTOP: {action}", "EXECUTE", 200, "127.0.0.1", "Tkinter Desktop Console App", "info", details]
            )
            conn.commit()
            conn.close()
        except Exception as e:
            print("Failed to record audit log:", e)

    # ── Tab 1: Motors Management ──────────────────────────────────────────────
    def create_motors_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Motors Catalog")
        
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        # Top Bar: Filters and Actions
        top_bar = ttk.Frame(tab)
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        ttk.Label(top_bar, text="Search:").grid(row=0, column=0, padx=5)
        self.motor_search_var = tk.StringVar()
        self.motor_search_var.trace_add("write", lambda *args: self.load_motors())
        search_entry = ttk.Entry(top_bar, textvariable=self.motor_search_var, width=25)
        search_entry.grid(row=0, column=1, padx=5)

        ttk.Label(top_bar, text="Category:").grid(row=0, column=2, padx=5)
        self.motor_cat_filter = ttk.Combobox(top_bar, state="readonly", width=15)
        self.motor_cat_filter.grid(row=0, column=3, padx=5)
        self.motor_cat_filter.bind("<<ComboboxSelected>>", lambda e: self.load_motors())

        # Buttons
        ttk.Button(top_bar, text="Add Motor", command=self.btn_add_motor).grid(row=0, column=4, padx=5)
        ttk.Button(top_bar, text="Edit Selected", command=self.btn_edit_motor).grid(row=0, column=5, padx=5)
        ttk.Button(top_bar, text="Delete Selected", style="Danger.TButton", command=self.btn_delete_motor).grid(row=0, column=6, padx=5)

        # Grid / Treeview
        tree_frame = ttk.Frame(tab)
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        self.motors_tree = ttk.Treeview(tree_frame, columns=("Name", "Company", "Thrust", "ESC", "Propeller", "Category"), show="headings")
        self.motors_tree.grid(row=0, column=0, sticky="nsew")
        
        self.motors_tree.heading("Name", text="Motor Name")
        self.motors_tree.heading("Company", text="Manufacturer")
        self.motors_tree.heading("Thrust", text="Max Thrust")
        self.motors_tree.heading("ESC", text="Recommended ESC")
        self.motors_tree.heading("Propeller", text="Recommended Propeller")
        self.motors_tree.heading("Category", text="Thrust Class")

        self.motors_tree.column("Name", width=220)
        self.motors_tree.column("Company", width=120)
        self.motors_tree.column("Thrust", width=100)
        self.motors_tree.column("ESC", width=150)
        self.motors_tree.column("Propeller", width=180)
        self.motors_tree.column("Category", width=120)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.motors_tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        self.motors_tree.configure(yscrollcommand=vsb.set)

        self.motors_data_map = {} # holds mapping of Treeview item to full record dict
        self.categories_list = [] # holds list of (id, name) categories
        self.load_categories_combo()
        self.load_motors()

    def load_categories_combo(self):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM public.categories ORDER BY name")
            self.categories_list = cursor.fetchall()
            conn.close()

            cat_names = ["All Categories"] + [c[1] for c in self.categories_list]
            self.motor_cat_filter['values'] = cat_names
            self.motor_cat_filter.current(0)
        except Exception as e:
            print("Failed to load categories for filter:", e)

    def load_motors(self):
        # Clear current grid
        self.motors_tree.delete(*self.motors_tree.get_children())
        self.motors_data_map.clear()

        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            sql = """
                SELECT m.id, m.motor_name, m.company, m.max_thrust, m.recommended_esc, m.recommended_propeller, 
                       m.link_motor, m.link_esc, m.link_propeller, m.custom_parameters, c.name, m.category_id 
                FROM public.motors m 
                LEFT JOIN public.categories c ON m.category_id = c.id
                WHERE 1=1
            """
            params = []
            
            # Apply Category Filter
            sel_cat_idx = self.motor_cat_filter.current()
            if sel_cat_idx > 0: # not "All Categories"
                cat_id = self.categories_list[sel_cat_idx - 1][0]
                sql += " AND m.category_id = %s"
                params.append(cat_id)
                
            # Apply Search Query
            search_val = self.motor_search_var.get().strip().lower()
            if search_val:
                sql += " AND (LOWER(m.motor_name) LIKE %s OR LOWER(m.company) LIKE %s)"
                params.append(f"%{search_val}%")
                
            sql += " ORDER BY m.company, m.motor_name"
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            
            for r in rows:
                item_id = self.motors_tree.insert("", "end", values=(r[1], r[2], r[3], r[4] or '', r[5] or '', r[10] or 'N/A'))
                self.motors_data_map[item_id] = {
                    'id': r[0],
                    'motor_name': r[1],
                    'company': r[2],
                    'max_thrust': r[3],
                    'recommended_esc': r[4] or '',
                    'recommended_propeller': r[5] or '',
                    'link_motor': r[6] or '',
                    'link_esc': r[7] or '',
                    'link_propeller': r[8] or '',
                    'custom_parameters': r[9] or {},
                    'category_name': r[10] or 'N/A',
                    'category_id': r[11]
                }
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load motors: {e}")

    def btn_add_motor(self):
        self.open_motor_editor(None)

    def btn_edit_motor(self):
        selected = self.motors_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a motor to edit.")
            return
        motor_data = self.motors_data_map[selected[0]]
        self.open_motor_editor(motor_data)

    def btn_delete_motor(self):
        selected = self.motors_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a motor to delete.")
            return
            
        motor_data = self.motors_data_map[selected[0]]
        confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to permanently delete \"{motor_data['motor_name']}\"?")
        if confirm:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM public.motors WHERE id = %s", [motor_data['id']])
                conn.commit()
                conn.close()
                self.log_event("DELETE_MOTOR", f"Deleted motor: {motor_data['motor_name']}")
                self.load_motors()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete motor: {e}")

    def open_motor_editor(self, motor_data=None):
        win = tk.Toplevel(self.root)
        win.title("Add Motor" if not motor_data else "Edit Motor")
        win.geometry("520x620")
        win.configure(bg=self.BG_DARK)
        
        # Grid weights
        win.grid_columnconfigure(1, weight=1)

        # Build Fields
        ttk.Label(win, text="Motor Specifications Form", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=15)

        ttk.Label(win, text="Category Class:").grid(row=1, column=0, sticky="w", padx=15, pady=5)
        cat_combo = ttk.Combobox(win, state="readonly")
        cat_combo.grid(row=1, column=1, sticky="ew", padx=15, pady=5)
        cat_names = [c[1] for c in self.categories_list]
        cat_combo['values'] = cat_names
        if motor_data:
            idx = next((i for i, c in enumerate(self.categories_list) if c[0] == motor_data['category_id']), -1)
            if idx != -1: cat_combo.current(idx)
        elif cat_names:
            cat_combo.current(0)

        # Helper to generate entry rows
        entries = {}
        fields = [
            ("motor_name", "Motor Name*", ""),
            ("company", "Manufacturer*", ""),
            ("max_thrust", "Max Thrust*", ""),
            ("recommended_esc", "Recommended ESC", ""),
            ("recommended_propeller", "Recommended Propeller", ""),
            ("link_motor", "Motor Spec Link", ""),
            ("link_esc", "ESC Spec Link", ""),
            ("link_propeller", "Propeller Spec Link", "")
        ]

        row_idx = 2
        for key, label, val in fields:
            ttk.Label(win, text=label).grid(row=row_idx, column=0, sticky="w", padx=15, pady=5)
            ent = ttk.Entry(win)
            ent.grid(row=row_idx, column=1, sticky="ew", padx=15, pady=5)
            if motor_data:
                ent.insert(0, motor_data.get(key, ""))
            entries[key] = ent
            row_idx += 1

        # Custom Parameters JSON
        ttk.Label(win, text="Custom Specs (JSON):").grid(row=row_idx, column=0, sticky="nw", padx=15, pady=5)
        json_txt = tk.Text(win, height=5, font=("Consolas", 10))
        json_txt.grid(row=row_idx, column=1, sticky="ew", padx=15, pady=5)
        if motor_data:
            json_txt.insert("1.0", json.dumps(motor_data.get("custom_parameters", {}), indent=2))
        else:
            json_txt.insert("1.0", "{\n  \n}")
        row_idx += 1

        def save_motor():
            cat_idx = cat_combo.current()
            if cat_idx == -1:
                messagebox.showerror("Validation Error", "Please select a valid category.")
                return
            cat_id = self.categories_list[cat_idx][0]
            
            # Simple field validations
            name = entries["motor_name"].get().strip()
            comp = entries["company"].get().strip()
            thru = entries["max_thrust"].get().strip()
            
            if not name or not comp or not thru:
                messagebox.showerror("Validation Error", "Please fill in all required (*) fields.")
                return

            # Validate links
            for link_key in ["link_motor", "link_esc", "link_propeller"]:
                link_val = entries[link_key].get().strip()
                if link_val and not (link_val.startswith("http://") or link_val.startswith("https://")):
                    messagebox.showerror("Validation Error", f"{link_key.replace('_',' ').title()} must be a valid URL starting with http:// or https://")
                    return

            # Validate custom json
            try:
                custom_json = json.loads(json_txt.get("1.0", tk.END).strip())
            except Exception as je:
                messagebox.showerror("JSON Parsing Error", f"Custom specifications field is not valid JSON:\n{je}")
                return

            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                
                if motor_data:
                    cursor.execute(
                        """
                        UPDATE public.motors 
                        SET category_id=%s, motor_name=%s, company=%s, max_thrust=%s, recommended_esc=%s, 
                            recommended_propeller=%s, link_motor=%s, link_esc=%s, link_propeller=%s, custom_parameters=%s 
                        WHERE id=%s
                        """,
                        [
                            cat_id, name, comp, thru,
                            entries["recommended_esc"].get().strip() or None,
                            entries["recommended_propeller"].get().strip() or None,
                            entries["link_motor"].get().strip() or None,
                            entries["link_esc"].get().strip() or None,
                            entries["link_propeller"].get().strip() or None,
                            json.dumps(custom_json),
                            motor_data['id']
                        ]
                    )
                    self.log_event("UPDATE_MOTOR", f"Updated motor spec: {name}")
                else:
                    cursor.execute(
                        """
                        INSERT INTO public.motors 
                        (category_id, motor_name, company, max_thrust, recommended_esc, recommended_propeller, link_motor, link_esc, link_propeller, custom_parameters)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        [
                            cat_id, name, comp, thru,
                            entries["recommended_esc"].get().strip() or None,
                            entries["recommended_propeller"].get().strip() or None,
                            entries["link_motor"].get().strip() or None,
                            entries["link_esc"].get().strip() or None,
                            entries["link_propeller"].get().strip() or None,
                            json.dumps(custom_json)
                        ]
                    )
                    self.log_event("CREATE_MOTOR", f"Created new motor spec: {name}")

                conn.commit()
                conn.close()
                win.destroy()
                self.load_motors()
            except Exception as ex:
                messagebox.showerror("Database Error", f"Operation failed: {ex}")

        # Buttons Panel
        btn_bar = ttk.Frame(win)
        btn_bar.grid(row=row_idx, column=0, columnspan=2, pady=20)
        ttk.Button(btn_bar, text="Save Spec", command=save_motor).grid(row=0, column=0, padx=10)
        ttk.Button(btn_bar, text="Cancel", command=win.destroy).grid(row=0, column=1, padx=10)

    # ── Tab 2: Categories Management ──────────────────────────────────────────
    def create_categories_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Categories Manager")
        
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        # Top Bar
        top_bar = ttk.Frame(tab)
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        ttk.Button(top_bar, text="Add Category", command=self.btn_add_cat).grid(row=0, column=0, padx=5)
        ttk.Button(top_bar, text="Edit Selected", command=self.btn_edit_cat).grid(row=0, column=1, padx=5)
        ttk.Button(top_bar, text="Delete Selected", style="Danger.TButton", command=self.btn_delete_cat).grid(row=0, column=2, padx=5)

        # Table View
        tree_frame = ttk.Frame(tab)
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        self.cats_tree = ttk.Treeview(tree_frame, columns=("ID", "Name", "Description"), show="headings")
        self.cats_tree.grid(row=0, column=0, sticky="nsew")

        self.cats_tree.heading("ID", text="Category ID")
        self.cats_tree.heading("Name", text="Thrust Class Name")
        self.cats_tree.heading("Description", text="Description")

        self.cats_tree.column("ID", width=250)
        self.cats_tree.column("Name", width=180)
        self.cats_tree.column("Description", width=550)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.cats_tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        self.cats_tree.configure(yscrollcommand=vsb.set)

        self.cats_data_map = {}
        self.load_categories()

    def load_categories(self):
        self.cats_tree.delete(*self.cats_tree.get_children())
        self.cats_data_map.clear()

        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, description FROM public.categories ORDER BY name")
            rows = cursor.fetchall()
            
            for r in rows:
                item_id = self.cats_tree.insert("", "end", values=(r[0], r[1], r[2] or ''))
                self.cats_data_map[item_id] = {
                    'id': r[0],
                    'name': r[1],
                    'description': r[2] or ''
                }
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load categories: {e}")

    def btn_add_cat(self):
        self.open_category_editor(None)

    def btn_edit_cat(self):
        selected = self.cats_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a category to edit.")
            return
        cat_data = self.cats_data_map[selected[0]]
        self.open_category_editor(cat_data)

    def btn_delete_cat(self):
        selected = self.cats_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a category to delete.")
            return
            
        cat_data = self.cats_data_map[selected[0]]
        confirm = messagebox.askyesno(
            "Confirm Delete", 
            f"Are you sure you want to permanently delete \"{cat_data['name']}\"?\nAll motors inside this category will also be deleted!"
        )
        if confirm:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM public.categories WHERE id = %s", [cat_data['id']])
                conn.commit()
                conn.close()
                self.log_event("DELETE_CATEGORY", f"Deleted category: {cat_data['name']}")
                self.load_categories()
                self.load_categories_combo()
                self.load_motors()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete category: {e}")

    def open_category_editor(self, cat_data=None):
        win = tk.Toplevel(self.root)
        win.title("Add Category" if not cat_data else "Edit Category")
        win.geometry("450x300")
        win.configure(bg=self.BG_DARK)
        win.grid_columnconfigure(1, weight=1)

        ttk.Label(win, text="Category Details Form", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=15)

        ttk.Label(win, text="Category Name*:").grid(row=1, column=0, sticky="w", padx=15, pady=5)
        name_ent = ttk.Entry(win)
        name_ent.grid(row=1, column=1, sticky="ew", padx=15, pady=5)
        if cat_data:
            name_ent.insert(0, cat_data["name"])

        ttk.Label(win, text="Description:").grid(row=2, column=0, sticky="nw", padx=15, pady=5)
        desc_txt = tk.Text(win, height=4, font=("Segoe UI", 10))
        desc_txt.grid(row=2, column=1, sticky="ew", padx=15, pady=5)
        if cat_data:
            desc_txt.insert("1.0", cat_data["description"])

        def save_category():
            name = name_ent.get().strip()
            desc = desc_txt.get("1.0", tk.END).strip()
            
            if not name:
                messagebox.showerror("Validation Error", "Category name is required.")
                return

            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                if cat_data:
                    cursor.execute("UPDATE public.categories SET name=%s, description=%s WHERE id=%s", [name, desc, cat_data['id']])
                    self.log_event("UPDATE_CATEGORY", f"Updated category: {name}")
                else:
                    cursor.execute("INSERT INTO public.categories (name, description) VALUES (%s, %s)", [name, desc])
                    self.log_event("CREATE_CATEGORY", f"Created category: {name}")
                conn.commit()
                conn.close()
                win.destroy()
                self.load_categories()
                self.load_categories_combo()
                self.load_motors()
            except Exception as e:
                messagebox.showerror("Database Error", f"Operation failed: {e}")

        # Buttons
        btn_bar = ttk.Frame(win)
        btn_bar.grid(row=3, column=0, columnspan=2, pady=15)
        ttk.Button(btn_bar, text="Save Category", command=save_category).grid(row=0, column=0, padx=10)
        ttk.Button(btn_bar, text="Cancel", command=win.destroy).grid(row=0, column=1, padx=10)

    # ── Tab 3: Users Management (Cognito Sync) ────────────────────────────────
    def create_users_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="User Accounts")
        
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        # Top Bar Actions
        top_bar = ttk.Frame(tab)
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        ttk.Button(top_bar, text="Create User Account", command=self.btn_add_user).grid(row=0, column=0, padx=5)
        ttk.Button(top_bar, text="Change Role", command=self.btn_change_role).grid(row=0, column=1, padx=5)
        ttk.Button(top_bar, text="Delete User Account", style="Danger.TButton", command=self.btn_delete_user).grid(row=0, column=2, padx=5)

        # Grid
        tree_frame = ttk.Frame(tab)
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        self.users_tree = ttk.Treeview(tree_frame, columns=("ID", "Email", "Role", "Created At"), show="headings")
        self.users_tree.grid(row=0, column=0, sticky="nsew")

        self.users_tree.heading("ID", text="User Pool UID (Sub)")
        self.users_tree.heading("Email", text="Account Email")
        self.users_tree.heading("Role", text="Access Role")
        self.users_tree.heading("Created At", text="Created At")

        self.users_tree.column("ID", width=260)
        self.users_tree.column("Email", width=220)
        self.users_tree.column("Role", width=120)
        self.users_tree.column("Created At", width=180)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.users_tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        self.users_tree.configure(yscrollcommand=vsb.set)

        self.users_data_map = {}
        self.load_users()

    def load_users(self):
        self.users_tree.delete(*self.users_tree.get_children())
        self.users_data_map.clear()

        try:
            # Query RDS database profiles
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, email, role, created_at FROM public.user_profiles ORDER BY email")
            rows = cursor.fetchall()
            conn.close()

            for r in rows:
                created_str = r[3].strftime("%Y-%m-%d %H:%M:%S") if isinstance(r[3], datetime) else str(r[3])
                item_id = self.users_tree.insert("", "end", values=(r[0], r[1], r[2], created_str))
                self.users_data_map[item_id] = {
                    'id': r[0],
                    'email': r[1],
                    'role': r[2],
                    'created_at': created_str
                }
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load user profiles from database: {e}")

    def btn_add_user(self):
        win = tk.Toplevel(self.root)
        win.title("Create User Account")
        win.geometry("420x350")
        win.configure(bg=self.BG_DARK)
        win.grid_columnconfigure(1, weight=1)

        ttk.Label(win, text="New Account Details", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=15)

        ttk.Label(win, text="Email Address*:").grid(row=1, column=0, sticky="w", padx=15, pady=5)
        email_ent = ttk.Entry(win)
        email_ent.grid(row=1, column=1, sticky="ew", padx=15, pady=5)

        ttk.Label(win, text="Password*:").grid(row=2, column=0, sticky="w", padx=15, pady=5)
        pass_ent = ttk.Entry(win, show="*")
        pass_ent.grid(row=2, column=1, sticky="ew", padx=15, pady=5)

        ttk.Label(win, text="User Role*:").grid(row=3, column=0, sticky="w", padx=15, pady=5)
        role_combo = ttk.Combobox(win, values=["guest", "intern", "admin"], state="readonly")
        role_combo.grid(row=3, column=1, sticky="ew", padx=15, pady=5)
        role_combo.current(0)

        def save_user():
            email = email_ent.get().strip()
            password = pass_ent.get().strip()
            role = role_combo.get()

            if not email or not password:
                messagebox.showerror("Validation Error", "All fields are required.")
                return

            try:
                # 1. Create Cognito user pool registration
                cognito_client = get_cognito_client()
                cog_username = str(uuid.uuid4())
                
                res = cognito_client.admin_create_user(
                    UserPoolId=COGNITO_USER_POOL_ID,
                    Username=cog_username,
                    UserAttributes=[
                        {'Name': 'email', 'Value': email},
                        {'Name': 'email_verified', 'Value': 'true'}
                    ],
                    TemporaryPassword=password,
                    MessageAction='SUPPRESS'
                )

                sub_uid = None
                for attr in res.get('User', {}).get('Attributes', []):
                    if attr['Name'] == 'sub':
                        sub_uid = attr['Value']
                        break

                if not sub_uid:
                    raise Exception("Cognito sub UUID not returned.")

                # Set password as permanent
                cognito_client.admin_set_user_password(
                    UserPoolId=COGNITO_USER_POOL_ID,
                    Username=cog_username,
                    Password=password,
                    Permanent=True
                )

                # 2. Save in database profile
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO public.user_profiles (id, email, role) VALUES (%s, %s, %s)",
                    [sub_uid, email, role]
                )
                conn.commit()
                conn.close()

                self.log_event("CREATE_USER", f"Created user account: {email} with role {role}")
                win.destroy()
                self.load_users()
                messagebox.showinfo("Success", f"User account \"{email}\" has been successfully created!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create user account:\n{e}")

        # Buttons
        btn_bar = ttk.Frame(win)
        btn_bar.grid(row=4, column=0, columnspan=2, pady=20)
        ttk.Button(btn_bar, text="Create Account", command=save_user).grid(row=0, column=0, padx=10)
        ttk.Button(btn_bar, text="Cancel", command=win.destroy).grid(row=0, column=1, padx=10)

    def btn_change_role(self):
        selected = self.users_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a user to update.")
            return
            
        user_data = self.users_data_map[selected[0]]
        
        win = tk.Toplevel(self.root)
        win.title("Change User Role")
        win.geometry("350x200")
        win.configure(bg=self.BG_DARK)
        win.grid_columnconfigure(1, weight=1)

        ttk.Label(win, text=f"Update role for: {user_data['email']}", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, columnspan=2, pady=15)

        ttk.Label(win, text="New Role:").grid(row=1, column=0, sticky="w", padx=15, pady=5)
        role_combo = ttk.Combobox(win, values=["guest", "intern", "admin"], state="readonly")
        role_combo.grid(row=1, column=1, sticky="ew", padx=15, pady=5)
        role_combo.set(user_data['role'])

        def save_role():
            new_role = role_combo.get()
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("UPDATE public.user_profiles SET role = %s WHERE id = %s", [new_role, user_data['id']])
                conn.commit()
                conn.close()

                self.log_event("UPDATE_USER_ROLE", f"Updated role of {user_data['email']} to {new_role}")
                win.destroy()
                self.load_users()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to change user role: {e}")

        btn_bar = ttk.Frame(win)
        btn_bar.grid(row=2, column=0, columnspan=2, pady=20)
        ttk.Button(btn_bar, text="Save Changes", command=save_role).grid(row=0, column=0, padx=10)
        ttk.Button(btn_bar, text="Cancel", command=win.destroy).grid(row=0, column=1, padx=10)

    def btn_delete_user(self):
        selected = self.users_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a user to delete.")
            return
            
        user_data = self.users_data_map[selected[0]]
        confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to permanently delete \"{user_data['email']}\" from database and Cognito User Pool?")
        if confirm:
            try:
                # 1. Delete from Database
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM public.user_onboarding WHERE user_id = %s", [user_data['id']])
                cursor.execute("DELETE FROM public.user_profiles WHERE id = %s", [user_data['id']])
                conn.commit()
                conn.close()

                # 2. Delete from Cognito User Pool
                cognito_client = get_cognito_client()
                try:
                    cognito_client.admin_delete_user(
                        UserPoolId=COGNITO_USER_POOL_ID,
                        Username=user_data['id']
                    )
                except cognito_client.exceptions.UserNotFoundException:
                    pass

                self.log_event("DELETE_USER", f"Permanently deleted account: {user_data['email']}")
                self.load_users()
                messagebox.showinfo("Success", f"Account \"{user_data['email']}\" deleted successfully.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete user: {e}")

    # ── Tab 4: Dynamic Spec Schema Customizer ──────────────────────────────────
    def create_schema_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Spec Schema")
        
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        # Top Bar Action
        top_bar = ttk.Frame(tab)
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        ttk.Button(top_bar, text="Add Schema Column", command=self.btn_add_schema_col).grid(row=0, column=0, padx=5)
        ttk.Button(top_bar, text="Delete Selected", style="Danger.TButton", command=self.btn_delete_schema_col).grid(row=0, column=1, padx=5)

        # Grid view
        tree_frame = ttk.Frame(tab)
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        self.schema_tree = ttk.Treeview(tree_frame, columns=("Key", "Name", "Type", "Unit"), show="headings")
        self.schema_tree.grid(row=0, column=0, sticky="nsew")

        self.schema_tree.heading("Key", text="JSON Field Key")
        self.schema_tree.heading("Name", text="Display Name")
        self.schema_tree.heading("Type", text="Value Type")
        self.schema_tree.heading("Unit", text="Unit (e.g. mm, V, g)")

        self.schema_tree.column("Key", width=220)
        self.schema_tree.column("Name", width=260)
        self.schema_tree.column("Type", width=120)
        self.schema_tree.column("Unit", width=100)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.schema_tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        self.schema_tree.configure(yscrollcommand=vsb.set)

        self.schema_data_map = {}
        self.load_schema()

    def load_schema(self):
        self.schema_tree.delete(*self.schema_tree.get_children())
        self.schema_data_map.clear()

        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, field_key, field_name, field_type, field_unit FROM public.custom_specs_schema ORDER BY field_name")
            rows = cursor.fetchall()
            conn.close()

            for r in rows:
                item_id = self.schema_tree.insert("", "end", values=(r[1], r[2], r[3], r[4] or '-'))
                self.schema_data_map[item_id] = {
                    'id': r[0],
                    'field_key': r[1],
                    'field_name': r[2],
                    'field_type': r[3],
                    'field_unit': r[4] or ''
                }
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load spec schema: {e}")

    def btn_add_schema_col(self):
        win = tk.Toplevel(self.root)
        win.title("Add Spec Column")
        win.geometry("400x320")
        win.configure(bg=self.BG_DARK)
        win.grid_columnconfigure(1, weight=1)

        ttk.Label(win, text="New Custom Spec Column", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=15)

        ttk.Label(win, text="Field Key (e.g. KV, weight)*:").grid(row=1, column=0, sticky="w", padx=15, pady=5)
        key_ent = ttk.Entry(win)
        key_ent.grid(row=1, column=1, sticky="ew", padx=15, pady=5)

        ttk.Label(win, text="Display Label Name*:").grid(row=2, column=0, sticky="w", padx=15, pady=5)
        name_ent = ttk.Entry(win)
        name_ent.grid(row=2, column=1, sticky="ew", padx=15, pady=5)

        ttk.Label(win, text="Data Type:").grid(row=3, column=0, sticky="w", padx=15, pady=5)
        type_combo = ttk.Combobox(win, values=["text", "number", "boolean"], state="readonly")
        type_combo.grid(row=3, column=1, sticky="ew", padx=15, pady=5)
        type_combo.current(0)

        ttk.Label(win, text="Unit (optional):").grid(row=4, column=0, sticky="w", padx=15, pady=5)
        unit_ent = ttk.Entry(win)
        unit_ent.grid(row=4, column=1, sticky="ew", padx=15, pady=5)

        def save_field():
            field_key = key_ent.get().strip().lower().replace(" ", "_")
            field_name = name_ent.get().strip()
            field_type = type_combo.get()
            field_unit = unit_ent.get().strip() or None

            if not field_key or not field_name:
                messagebox.showerror("Validation Error", "Field key and display name are required.")
                return

            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO public.custom_specs_schema (field_key, field_name, field_type, field_unit) VALUES (%s, %s, %s, %s)",
                    [field_key, field_name, field_type, field_unit]
                )
                conn.commit()
                conn.close()

                self.log_event("CREATE_SCHEMA_COLUMN", f"Created custom spec field: {field_name} ({field_key})")
                win.destroy()
                self.load_schema()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add schema column: {e}")

        # Buttons
        btn_bar = ttk.Frame(win)
        btn_bar.grid(row=5, column=0, columnspan=2, pady=15)
        ttk.Button(btn_bar, text="Add Column", command=save_field).grid(row=0, column=0, padx=10)
        ttk.Button(btn_bar, text="Cancel", command=win.destroy).grid(row=0, column=1, padx=10)

    def btn_delete_schema_col(self):
        selected = self.schema_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a schema column to delete.")
            return
            
        field_data = self.schema_data_map[selected[0]]
        confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to permanently delete column \"{field_data['field_name']}\"?")
        if confirm:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM public.custom_specs_schema WHERE id = %s", [field_data['id']])
                conn.commit()
                conn.close()

                self.log_event("DELETE_SCHEMA_COLUMN", f"Deleted custom spec field: {field_data['field_name']}")
                self.load_schema()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete schema column: {e}")

    # ── Tab 5: Security Operations Logs ────────────────────────────────────────
    def create_logs_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Security Logs")
        
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        # Top Bar
        top_bar = ttk.Frame(tab)
        top_bar.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        
        ttk.Button(top_bar, text="Refresh Logs", command=self.load_logs).grid(row=0, column=0, padx=5)
        
        ttk.Label(top_bar, text="Filter Risk:").grid(row=0, column=1, padx=5)
        self.log_risk_filter = ttk.Combobox(top_bar, values=["All", "info", "warning", "suspicious"], state="readonly", width=12)
        self.log_risk_filter.grid(row=0, column=2, padx=5)
        self.log_risk_filter.current(0)
        self.log_risk_filter.bind("<<ComboboxSelected>>", lambda e: self.load_logs())

        # Grid view
        tree_frame = ttk.Frame(tab)
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        self.logs_tree = ttk.Treeview(tree_frame, columns=("Time", "User", "Action", "Status", "IP", "Risk", "Details"), show="headings")
        self.logs_tree.grid(row=0, column=0, sticky="nsew")

        self.logs_tree.heading("Time", text="Timestamp")
        self.logs_tree.heading("User", text="Operator User")
        self.logs_tree.heading("Action", text="Operation Action")
        self.logs_tree.heading("Status", text="Status")
        self.logs_tree.heading("IP", text="IP Address")
        self.logs_tree.heading("Risk", text="Risk Level")
        self.logs_tree.heading("Details", text="Operation Description Details")

        self.logs_tree.column("Time", width=160)
        self.logs_tree.column("User", width=160)
        self.logs_tree.column("Action", width=160)
        self.logs_tree.column("Status", width=60)
        self.logs_tree.column("IP", width=110)
        self.logs_tree.column("Risk", width=80)
        self.logs_tree.column("Details", width=340)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.logs_tree.yview)
        vsb.grid(row=0, column=1, sticky="ns")
        self.logs_tree.configure(yscrollcommand=vsb.set)

        self.load_logs()

    def load_logs(self):
        self.logs_tree.delete(*self.logs_tree.get_children())

        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            sql = "SELECT timestamp, email, route, method, status, ip_address, risk_level, details FROM public.audit_logs WHERE 1=1"
            params = []
            
            # Risk Filter
            risk_sel = self.log_risk_filter.get()
            if risk_sel != "All":
                sql += " AND risk_level = %s"
                params.append(risk_sel)
                
            sql += " ORDER BY timestamp DESC LIMIT 200"
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            
            for r in rows:
                time_str = r[0].strftime("%Y-%m-%d %H:%M:%S") if isinstance(r[0], datetime) else str(r[0])
                self.logs_tree.insert("", "end", values=(time_str, r[1], r[2], r[4], r[5], r[6], r[7]))
            conn.close()
        except Exception as e:
            # Table might not exist or empty
            print("Logs loading failed:", e)

    # ── Tab 6: Excel / CSV Bulk Data Imports & Exports ───────────────────────
    def create_data_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Bulk Data Actions")
        
        tab.grid_columnconfigure(0, weight=1)
        
        # Guide Panel
        guide_frame = ttk.LabelFrame(tab, text=" Curation Pipeline Operations ", padding=20)
        guide_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=20)
        guide_frame.grid_columnconfigure(0, weight=1)
        
        guide_text = (
            "ThrustVault Admin Console Pipeline Utilities:\n\n"
            "1. Bulk Seeding: Import and override categories & motors catalog directly using 'Motor List.xlsx'.\n"
            "   This action deletes all existing database entries and resets standard specs sheets.\n\n"
            "2. Table Exporting: Extract categories and drone specs database tables into spreadsheet templates."
        )
        ttk.Label(guide_frame, text=guide_text, justify="left", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w")

        # Seeding Actions Group
        import_frame = ttk.LabelFrame(tab, text=" Import Actions ", padding=20)
        import_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=10)
        
        ttk.Button(import_frame, text="Bulk Import Excel File (Reseed DB)", command=self.btn_bulk_import_reseed).grid(row=0, column=0, padx=10, pady=5)

        # Export Actions Group
        export_frame = ttk.LabelFrame(tab, text=" Export Actions ", padding=20)
        export_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=10)
        
        ttk.Button(export_frame, text="Export Motors to Excel (XLSX)", command=self.btn_export_xlsx).grid(row=0, column=0, padx=10, pady=5)
        ttk.Button(export_frame, text="Export Motors to CSV", command=self.btn_export_csv).grid(row=0, column=1, padx=10, pady=5)

    def btn_bulk_import_reseed(self):
        file_path = filedialog.askopenfilename(
            title="Select Motor List Excel File",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path:
            return

        confirm = messagebox.askyesno(
            "CRITICAL RESET WARNING",
            "WARNING: This action will completely erase all current categories and motors in the database, and replace them with the parsed rows from the Excel sheets.\n\nAre you sure you want to perform this operation?"
        )
        if not confirm:
            return

        try:
            # Parse sheets
            wb = openpyxl.load_workbook(file_path)
            sheet_map = {
                '2kg':  {'name': '1-2 kg', 'description': 'Standard multirotor and inspection drones (typically 3S – 4S LiPo)'},
                '5kg':  {'name': '3-5 kg', 'description': 'Mid-range commercial multirotors and survey drones (typically 4S – 6S)'},
                '10kg': {'name': '8-10 kg', 'description': 'Heavy-lift inspection and payload drones (typically 6S – 12S)'},
                '20kg': {'name': '18-22 kg', 'description': 'Industrial mapping, agricultural and large-frame drones (typically 12S – 14S)'},
                '50kg': {'name': '45-55 kg', 'description': 'High-payload cargo and heavy agricultural drones (typically 14S – 24S)'},
            }

            data = {}
            for sheet_name, cat_info in sheet_map.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
                motors = []
                for row in rows[2:]:   # skip title row + header row
                    motor = str(row[0]).strip() if row[0] else ''
                    if not motor or motor in ('Motor', 'MOTOR', ' '):
                        continue
                    company = str(row[1]).strip() if row[1] else ''
                    thrust  = str(row[2]).strip() if row[2] else ''
                    esc     = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    prop    = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    link_m  = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                    link_e  = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                    link_p  = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                    
                    # Clean URLs to pass constraints
                    def check_url(u):
                        if not u: return None
                        u = u.strip()
                        if u.lower().startswith('http://') or u.lower().startswith('https://'):
                            return u
                        if u.lower().startswith('www.'):
                            return 'https://' + u
                        return None

                    motors.append({
                        'motor_name': motor,
                        'company': company,
                        'max_thrust': thrust,
                        'recommended_esc': esc or None,
                        'recommended_propeller': prop.replace('\n', ' ').strip() if prop else None,
                        'link_motor': check_url(link_m),
                        'link_esc': check_url(link_e),
                        'link_propeller': check_url(link_p),
                    })
                data[cat_info['name']] = {
                    'description': cat_info['description'],
                    'motors': motors
                }

            # Transaction DB Writes
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM public.motors")
            cursor.execute("DELETE FROM public.categories")

            total_motors = 0
            for cat_name, cat_data in data.items():
                cursor.execute(
                    "INSERT INTO public.categories (name, description) VALUES (%s, %s) RETURNING id",
                    [cat_name, cat_data['description']]
                )
                cat_id = cursor.fetchone()[0]

                for m in cat_data['motors']:
                    cursor.execute(
                        """
                        INSERT INTO public.motors 
                        (category_id, motor_name, company, max_thrust, recommended_esc, recommended_propeller, link_motor, link_esc, link_propeller)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        [
                            cat_id, m['motor_name'], m['company'], m['max_thrust'],
                            m['recommended_esc'], m['recommended_propeller'],
                            m['link_motor'], m['link_esc'], m['link_propeller']
                        ]
                    )
                    total_motors += 1

            conn.commit()
            conn.close()

            self.log_event("BULK_IMPORT", f"Reseeded database: {len(data)} categories, {total_motors} motors")
            self.load_categories()
            self.load_categories_combo()
            self.load_motors()
            messagebox.showinfo("Success", f"Database re-seeded successfully!\nLoaded {len(data)} categories and {total_motors} motors.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to bulk-import and reseed:\n{e}")

    def btn_export_xlsx(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Export File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Motors Export"
            
            headers = ["Category", "Motor Name", "Manufacturer", "Max Thrust", "Recommended ESC", "Recommended Propeller", "Motor Link", "ESC Link", "Propeller Link"]
            ws.append(headers)

            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT c.name, m.motor_name, m.company, m.max_thrust, m.recommended_esc, m.recommended_propeller, 
                       m.link_motor, m.link_esc, m.link_propeller 
                FROM public.motors m
                LEFT JOIN public.categories c ON m.category_id = c.id
                ORDER BY c.name, m.company, m.motor_name
                """
            )
            rows = cursor.fetchall()
            for r in rows:
                ws.append([r[0] or 'N/A', r[1], r[2], r[3], r[4] or '', r[5] or '', r[6] or '', r[7] or '', r[8] or ''])
            
            wb.save(file_path)
            conn.close()
            self.log_event("EXPORT_EXCEL", f"Exported motors catalog to Excel: {file_path}")
            messagebox.showinfo("Success", "Catalog exported to Excel spreadsheet successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}")

    def btn_export_csv(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Export File",
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")]
        )
        if not file_path:
            return

        try:
            import csv
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT c.name, m.motor_name, m.company, m.max_thrust, m.recommended_esc, m.recommended_propeller 
                FROM public.motors m
                LEFT JOIN public.categories c ON m.category_id = c.id
                ORDER BY c.name, m.company, m.motor_name
                """
            )
            rows = cursor.fetchall()
            
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Category", "Motor Name", "Manufacturer", "Max Thrust", "Recommended ESC", "Recommended Propeller"])
                for r in rows:
                    writer.writerow([r[0] or 'N/A', r[1], r[2], r[3], r[4] or '', r[5] or ''])
            conn.close()
            self.log_event("EXPORT_CSV", f"Exported motors catalog to CSV: {file_path}")
            messagebox.showinfo("Success", "Catalog exported to CSV file successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export CSV: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ThrustVaultAdminApp(root)
    root.mainloop()
