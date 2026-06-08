"""
seed_motors.py — Clears all existing motors/categories and re-seeds from Motor List.xlsx
Run:  python seed_motors.py
"""

import urllib.request
import json
import os
import openpyxl

# ── Load .env ─────────────────────────────────────────────────────────────────
def load_env(path='.env'):
    env = {}
    if os.path.exists(path):
        for line in open(path, encoding='utf-8'):
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                env[k.strip()] = v.strip()
    return env

env = load_env()
SUPABASE_URL      = env.get('SUPABASE_URL', os.environ.get('SUPABASE_URL', ''))
SERVICE_ROLE_KEY  = env.get('service_role', os.environ.get('service_role', ''))

if not SUPABASE_URL or not SERVICE_ROLE_KEY:
    print('ERROR: SUPABASE_URL or service_role not found in .env')
    exit(1)

# ── Supabase REST helper ──────────────────────────────────────────────────────
def sb(table, method='GET', payload=None, params=''):
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    if params:
        url += '?' + params
    headers = {
        'apikey':        SERVICE_ROLE_KEY,
        'Authorization': f'Bearer {SERVICE_ROLE_KEY}',
        'Content-Type':  'application/json',
        'Prefer':        'return=representation'
    }
    data = json.dumps(payload).encode() if payload is not None else None
    req  = urllib.request.Request(url, headers=headers, method=method, data=data)
    try:
        with urllib.request.urlopen(req) as res:
            body = res.read().decode()
            return json.loads(body) if body else []
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f'  HTTP {e.code} on {method} {table}: {body}')
        raise

# ── Read Excel ────────────────────────────────────────────────────────────────
def read_motors_from_excel(path):
    wb = openpyxl.load_workbook(path)
    sheet_map = {
        '2kg':  {'name': '1-2 kg', 'description': 'Standard multirotor and inspection drones (typically 3S – 4S LiPo)'},
        '5kg':  {'name': '3-5 kg', 'description': 'Mid-range commercial multirotors and survey drones (typically 4S – 6S)'},
        '10kg': {'name': '8-10 kg', 'description': 'Heavy-lift inspection and payload drones (typically 6S – 12S)'},
        '20kg': {'name': '18-22 kg', 'description': 'Industrial mapping, agricultural and large-frame drones (typically 12S – 14S)'},
        '50kg': {'name': '45-55 kg', 'description': 'High-payload cargo and heavy agricultural drones (typically 14S – 24S)'},
    }

    result = {}
    for sheet_name, cat_info in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws   = wb[sheet_name]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        motors = []
        for row in rows[2:]:   # skip title row + header row
            motor = str(row[0]).strip() if row[0] else ''
            if not motor or motor in ('Motor', 'MOTOR', ' '):
                continue
            company = str(row[1]).strip() if row[1] else ''
            thrust  = str(row[2]).strip() if row[2] else ''
            esc     = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            prop    = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            link_m  = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            link_e  = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            link_p  = str(row[7]).strip() if len(row) > 7 and row[7] else ''
            # clean newlines in prop names
            prop = prop.replace('\n', ' ').strip()
            motors.append({
                'motor_name':            motor,
                'company':               company,
                'max_thrust':            thrust,
                'recommended_esc':       esc   or None,
                'recommended_propeller': prop  or None,
                'link_motor':            link_m or None,
                'link_esc':              link_e or None,
                'link_propeller':        link_p or None,
            })
        result[cat_info['name']] = {
            'description': cat_info['description'],
            'motors':      motors
        }
        print(f'  Parsed sheet [{sheet_name}] → category "{cat_info["name"]}": {len(motors)} motors')
    return result

# ── Main ──────────────────────────────────────────────────────────────────────
def seed():
    print('=' * 60)
    print('ThrustVault Motor Seed Script')
    print('=' * 60)

    # 1. Delete all existing motors (cascade delete from test_runs handled by FK)
    print('\n[1/4] Deleting all existing motors...')
    sb('motors', method='DELETE', params='id=neq.00000000-0000-0000-0000-000000000000')
    print('  Done.')

    # 2. Delete all existing categories
    print('[2/4] Deleting all existing categories...')
    sb('categories', method='DELETE', params='id=neq.00000000-0000-0000-0000-000000000000')
    print('  Done.')

    # 3. Parse Excel
    print('\n[3/4] Reading Motor List.xlsx...')
    excel_path = r'd:\motor data\Motor List.xlsx'
    data = read_motors_from_excel(excel_path)

    # 4. Insert categories then motors
    print('\n[4/4] Inserting new categories and motors...')
    total_motors = 0
    for cat_name, cat_data in data.items():
        # Insert category
        cat_res = sb('categories', method='POST', payload=[{
            'name':        cat_name,
            'description': cat_data['description']
        }])
        cat_id = cat_res[0]['id']
        print(f'\n  ✓ Category: "{cat_name}" (id: {cat_id[:8]}...)')

        # Attach category_id and insert motors in batches of 20
        motors = cat_data['motors']
        for m in motors:
            m['category_id'] = cat_id

        batch_size = 20
        for i in range(0, len(motors), batch_size):
            batch = motors[i:i + batch_size]
            sb('motors', method='POST', payload=batch)
            print(f'    Inserted motors {i+1}–{i+len(batch)} of {len(motors)}')

        total_motors += len(motors)

    print('\n' + '=' * 60)
    print(f'Seeding complete!')
    print(f'  Categories: {len(data)}')
    print(f'  Total motors: {total_motors}')
    print('=' * 60)

if __name__ == '__main__':
    seed()
